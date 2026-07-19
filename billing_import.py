"""
Historical Bill Number List workbook import (SPEC 6.2).

The workbook has 80+ monthly tabs with a consistent-enough structure:
fee note no., entity, client code, net, issued by, issued on. This
parser is deliberately tolerant — it finds the header row per sheet by
fuzzy-matching column names, reports everything it skipped, and nothing
is committed until the preview is accepted in the UI.

Built against the SPEC's description of the workbook; expect one round
of tuning when the real file is first uploaded.
"""

import datetime as dt
import io
import re

from openpyxl import load_workbook

# fuzzy header -> canonical field
HEADER_PATTERNS = {
    "fee_note_no": re.compile(r"(fee\s*note|bill|invoice)\s*(no|number|#)?", re.I),
    "entity": re.compile(r"entity|company|aa\s*/\s*cw", re.I),
    "client_code": re.compile(r"client\s*code|code", re.I),
    "client_name": re.compile(r"client(\s*name)?$|^name$", re.I),
    "net": re.compile(r"^net|net\s*(fee|amount)|amount\s*\(?net", re.I),
    "issued_by": re.compile(r"issued\s*by|raised\s*by|initials|by$", re.I),
    "issued_on": re.compile(r"issued\s*(on|date)|^date", re.I),
}
REQUIRED = ("fee_note_no", "net")


def _match_headers(cells) -> dict:
    """Map column index -> field if this row looks like a header row."""
    mapping = {}
    for idx, value in enumerate(cells):
        if not isinstance(value, str):
            continue
        text = value.strip()
        if not text:
            continue
        for field, pattern in HEADER_PATTERNS.items():
            if field not in mapping.values() and pattern.search(text):
                mapping[idx] = field
                break
    found = set(mapping.values())
    return mapping if all(f in found for f in REQUIRED) else {}


def _parse_date(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d %b %Y", "%d %B %Y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _parse_net(value):
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, str):
        text = value.replace("£", "").replace(",", "").strip()
        try:
            return round(float(text), 2)
        except ValueError:
            return None
    return None


def _norm_entity(value):
    if not isinstance(value, str):
        return None
    text = value.strip().upper()
    if text in ("AA", "A"):
        return "AA"
    if text in ("CW", "C"):
        return "CW"
    return None


def parse_workbook(file_bytes: bytes) -> dict:
    """Parse every sheet. Returns {rows, issues, sheets_parsed,
    sheets_skipped} — nothing is written to the DB here."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    rows, issues, parsed, skipped = [], [], [], []

    for ws in wb.worksheets:
        sheet_monthly = 1 if re.search(r"monthly", ws.title, re.I) and \
            not re.search(r"non", ws.title, re.I) else None
        mapping = {}
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15,
                                             values_only=True), start=1):
            mapping = _match_headers(list(row))
            if mapping:
                header_row_idx = i
                break
        if not mapping:
            skipped.append(ws.title)
            continue

        sheet_count = 0
        for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1,
                                             values_only=True),
                                start=header_row_idx + 1):
            values = {field: row[idx] if idx < len(row) else None
                      for idx, field in mapping.items()}
            fee_no = values.get("fee_note_no")
            net = _parse_net(values.get("net"))
            if fee_no is None and net is None:
                continue  # blank/total row
            if net is None:
                issues.append(f"{ws.title} row {i}: no net amount "
                              f"(fee note {fee_no}) — skipped")
                continue
            if fee_no is None:
                issues.append(f"{ws.title} row {i}: net £{net} with no fee "
                              "note number — skipped")
                continue
            monthly = sheet_monthly
            rows.append({
                "fee_note_no": str(fee_no).strip(),
                "entity": _norm_entity(values.get("entity")),
                "client_code": (str(values["client_code"]).strip()
                                if values.get("client_code") is not None else None),
                "client_name": (str(values["client_name"]).strip()
                                if values.get("client_name") is not None else None),
                "net": net,
                "issued_by": (str(values["issued_by"]).strip().upper()
                              if values.get("issued_by") is not None else None),
                "issued_on": _parse_date(values.get("issued_on")),
                "monthly": monthly,
                "source_tab": ws.title,
            })
            sheet_count += 1
        parsed.append(f"{ws.title}: {sheet_count} rows")

    undated = sum(1 for r in rows if not r["issued_on"])
    if undated:
        issues.append(f"{undated} row(s) have no parseable issued-on date — "
                      "they will import but won't appear in monthly views.")
    return {"rows": rows, "issues": issues,
            "sheets_parsed": parsed, "sheets_skipped": skipped}
